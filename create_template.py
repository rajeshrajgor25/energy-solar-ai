from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Solar Load Calculator"

# ── Styles ──────────────────────────────────────────────────────────────────
ORANGE   = "FF6B00"
YELLOW   = "FFF3CD"
LBLUE    = "E8F4FD"
WHITE    = "FFFFFF"
DGRAY    = "2C3E50"
LGRAY    = "F8F9FA"
GREEN    = "27AE60"

def hdr(text, row, col, bg=ORANGE, fg="FFFFFF", bold=True, size=11, span=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    return c

def lbl(text, row, col, bg=LGRAY, bold=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold, color=DGRAY, size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c

def inp(value, row, col, fmt=None, bg=WHITE):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(color="0000CC", size=10, bold=True)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        c.number_format = fmt
    # Highlight input cells with yellow border
    thin = Side(style="medium", color="FF6B00")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

def calc(formula, row, col, fmt=None, bg=LBLUE):
    c = ws.cell(row=row, column=col, value=formula)
    c.font = Font(color=DGRAY, size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        c.number_format = fmt
    return c

thin_side = Side(style="thin", color="CCCCCC")
thin_border = Border(left=thin_side, right=thin_side,
                     top=thin_side, bottom=thin_side)

# ── Column widths ────────────────────────────────────────────────────────────
col_widths = [2, 32, 22, 22, 18, 2]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Row heights
for r in range(1, 60):
    ws.row_dimensions[r].height = 22

# ── TITLE ───────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 40
ws.merge_cells("A1:F1")
t = ws["A1"]
t.value = "⚡ ENERGYBAE — Solar Load Calculator"
t.font = Font(bold=True, color=WHITE, size=16)
t.fill = PatternFill("solid", fgColor=ORANGE)
t.alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[2].height = 15
ws.merge_cells("A2:F2")
sub = ws["A2"]
sub.value = "Electricity Bill → Solar System Sizing Automation"
sub.font = Font(italic=True, color=DGRAY, size=10)
sub.fill = PatternFill("solid", fgColor=YELLOW)
sub.alignment = Alignment(horizontal="center", vertical="center")

# ── SECTION 1: Customer & Bill Info ─────────────────────────────────────────
hdr("SECTION 1 — Customer & Bill Information", 4, 2, span=3)

rows_s1 = [
    (5,  "Consumer Name",           "B5",  None),
    (6,  "Consumer Number",         "B6",  None),
    (7,  "Meter Number",            "B7",  None),
    (8,  "Billing Month",           "B8",  None),
    (9,  "Billing Address",         "B9",  None),
    (10, "MSEDCL Division",        "B10",  None),
]
for row, label, ref, fmt in rows_s1:
    lbl(label, row, 2)
    inp("", row, 3, fmt, WHITE)
    # col 4 empty
    ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=LGRAY)

# ── SECTION 2: Consumption Data ─────────────────────────────────────────────
hdr("SECTION 2 — Consumption & Tariff Data", 12, 2, span=3)

rows_s2 = [
    (13, "Units Consumed (kWh/month)",     "C13", "#,##0.00"),
    (14, "Sanctioned Load (kW)",           "C14", "#,##0.00"),
    (15, "Connected Load (kW)",            "C15", "#,##0.00"),
    (16, "Contract Demand (kVA)",          "C16", "#,##0.00"),
    (17, "Tariff Category",               "C17", None),
    (18, "Per Unit Rate (₹/kWh)",         "C18", "₹#,##0.00"),
    (19, "Fixed Charges (₹/month)",       "C19", "₹#,##0.00"),
    (20, "Total Monthly Bill (₹)",        "C20", "₹#,##0.00"),
    (21, "Annual Units Consumed (kWh)",   "C21", "#,##0.00"),
]
for row, label, ref, fmt in rows_s2:
    lbl(label, row, 2)
    inp("", row, 3, fmt, WHITE)
    ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=LGRAY)

# Annual auto-calc note
ws.cell(row=21, column=3).value = "=IF(C13>0,C13*12,\"\")"
ws.cell(row=21, column=3).font = Font(color=DGRAY, size=10)
ws.cell(row=21, column=3).fill = PatternFill("solid", fgColor=LBLUE)
ws.cell(row=21, column=3).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row=21, column=3).number_format = "#,##0.00"

# ── SECTION 3: Solar Assumptions ────────────────────────────────────────────
hdr("SECTION 3 — Solar System Assumptions", 23, 2, span=3)

rows_s3 = [
    (24, "Peak Sun Hours (hrs/day)",         "C24", "#,##0.0",  4.5),
    (25, "System Efficiency (%)",            "C25", "0.0%",     0.8),
    (26, "Solar Panel Wattage (Wp)",         "C26", "#,##0",    540),
    (27, "Cost per kWp (₹)",                "C27", "₹#,##0",   55000),
    (28, "Subsidy (%)",                      "C28", "0.0%",     0.30),
    (29, "Annual Degradation Rate (%)",      "C29", "0.0%",     0.005),
    (30, "System Lifespan (Years)",          "C30", "#,##0",    25),
]
for row, label, ref, fmt, default in rows_s3:
    lbl(label, row, 2)
    inp(default, row, 3, fmt, YELLOW)
    ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=LGRAY)

# ── SECTION 4: Solar Sizing Results ─────────────────────────────────────────
hdr("SECTION 4 — Solar System Sizing Results", 32, 2, span=3)

# Daily energy need
lbl("Daily Energy Requirement (kWh/day)", 33, 2, LGRAY)
calc("=IF(C13>0,C13/30,\"\")", 33, 3, "#,##0.00")

# Required system size
lbl("Required Solar Capacity (kWp)", 34, 2, LGRAY)
calc("=IF(AND(C24>0,C25>0,C13>0),(C13/30)/(C24*C25),\"\")", 34, 3, "#,##0.00")

# Recommended system size (round up to nearest 0.5)
lbl("Recommended System Size (kWp)", 35, 2, LGRAY)
calc("=IF(C34<>\"\",CEILING(C34,0.5),\"\")", 35, 3, "#,##0.00")

# Number of panels
lbl("Number of Solar Panels Required", 36, 2, LGRAY)
calc("=IF(AND(C35<>\"\",C26>0),CEILING((C35*1000)/C26,1),\"\")", 36, 3, "#,##0")

# Rooftop area
lbl("Rooftop Area Required (sq. ft.)", 37, 2, LGRAY)
calc("=IF(C35<>\"\",C35*100,\"\")", 37, 3, "#,##0")

# Annual generation
lbl("Estimated Annual Generation (kWh)", 38, 2, LGRAY)
calc("=IF(AND(C35<>\"\",C24>0),C35*C24*365*C25,\"\")", 38, 3, "#,##0.00")

# ── SECTION 5: Financial Analysis ───────────────────────────────────────────
hdr("SECTION 5 — Financial Analysis & ROI", 40, 2, span=3)

lbl("Gross System Cost (₹)", 41, 2, LGRAY)
calc("=IF(C35<>\"\",C35*C27,\"\")", 41, 3, "₹#,##0")

lbl("Subsidy Amount (₹)", 42, 2, LGRAY)
calc("=IF(C41<>\"\",C41*C28,\"\")", 42, 3, "₹#,##0")

lbl("Net System Cost After Subsidy (₹)", 43, 2, LGRAY)
calc("=IF(AND(C41<>\"\",C42<>\"\"),C41-C42,\"\")", 43, 3, "₹#,##0")

lbl("Annual Savings (₹)", 44, 2, LGRAY)
calc("=IF(AND(C38<>\"\",C18>0),C38*C18,\"\")", 44, 3, "₹#,##0")

lbl("Payback Period (Years)", 45, 2, LGRAY)
calc("=IF(AND(C43<>\"\",C44>0),C43/C44,\"\")", 45, 3, "#,##0.0")

lbl("25-Year Total Savings (₹)", 46, 2, LGRAY)
calc("=IF(AND(C44<>\"\",C43<>\"\"),C44*C30-C43,\"\")", 46, 3, "₹#,##0")

lbl("CO₂ Offset (Tonnes/Year)", 47, 2, LGRAY)
calc("=IF(C38<>\"\",C38*0.82/1000,\"\")", 47, 3, "#,##0.00")

# ── SECTION 6: Summary Box ──────────────────────────────────────────────────
ws.row_dimensions[49].height = 30
hdr("✅ RECOMMENDATION SUMMARY", 49, 2, bg=GREEN, span=3)

summary_items = [
    (50, "Customer",               "=IF(B5<>\"\",B5,\"—\")"),
    (51, "Billing Month",          "=IF(B8<>\"\",B8,\"—\")"),
    (52, "Monthly Units (kWh)",    "=IF(C13>0,C13,\"—\")"),
    (53, "Recommended Size (kWp)", "=IF(C35<>\"\",C35,\"—\")"),
    (54, "No. of Panels",          "=IF(C36<>\"\",C36,\"—\")"),
    (55, "Net Cost (₹)",           "=IF(C43<>\"\",C43,\"—\")"),
    (56, "Payback Period (Yrs)",   "=IF(C45<>\"\",C45,\"—\")"),
    (57, "Annual Savings (₹)",     "=IF(C44<>\"\",C44,\"—\")"),
]
for row, label, formula in summary_items:
    lbl(label, row, 2, LGRAY)
    c = ws.cell(row=row, column=3, value=formula)
    c.font = Font(bold=True, color=DGRAY, size=10)
    c.fill = PatternFill("solid", fgColor=LBLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")

# Apply thin borders to data range
for row in range(4, 58):
    for col in range(2, 5):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border

# ── Footer ───────────────────────────────────────────────────────────────────
ws.merge_cells("A59:F59")
foot = ws["A59"]
foot.value = "Generated by EnergyBae Solar Load Calculator AI | energybae.co@gmail.com"
foot.font = Font(italic=True, color="888888", size=9)
foot.alignment = Alignment(horizontal="center")

path = "/home/claude/solar-load-ai/templates/solar_template.xlsx"
wb.save(path)
print(f"Template saved: {path}")
