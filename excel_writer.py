import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook


TEMPLATE_PATH = Path(__file__).parent / "templates" / "solar_template.xlsx"
OUTPUT_DIR    = Path(__file__).parent / "outputs"

FIELD_CELL_MAP = {
    "consumer_name":     "C5",
    "consumer_number":   "C6",
    "meter_number":      "C7",
    "billing_month":     "C8",
    "billing_address":   "C9",
    "division":          "C10",
    "units_consumed":    "C13",
    "sanctioned_load":   "C14",
    "connected_load":    "C15",
    "contract_demand":   "C16",
    "tariff_category":   "C17",
    "per_unit_rate":     "C18",
    "fixed_charges":     "C19",
    "total_monthly_bill": "C20",
}


def fill_excel(bill_data: dict, solar_assumptions: dict = None) -> Path:
    """
    Fill the Excel template with extracted bill data + optional assumption overrides.
    Returns path to the generated output file.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    consumer = bill_data.get("consumer_name", "customer")
    safe_name = "".join(c for c in str(consumer) if c.isalnum() or c in " _-")[:20].strip()
    out_path = OUTPUT_DIR / f"Solar_Report_{safe_name}_{ts}.xlsx"
    
    # Copy template
    shutil.copy2(TEMPLATE_PATH, out_path)
    
    # Load workbook (keep formulas intact)
    wb = load_workbook(out_path)
    ws = wb.active
    
    # ── Fill input cells ─────────────────────────────────────────────────────
    for field, cell_ref in FIELD_CELL_MAP.items():
        value = bill_data.get(field)
        if value is not None and value != "null" and value != "":
            try:
                # Numeric fields
                if field in ("units_consumed", "sanctioned_load", "connected_load",
                             "contract_demand", "per_unit_rate", "fixed_charges",
                             "total_monthly_bill"):
                    ws[cell_ref] = float(value) if value else None
                else:
                    ws[cell_ref] = str(value)
            except (ValueError, TypeError):
                ws[cell_ref] = value
    
    # ── Override solar assumptions if provided ───────────────────────────────
    if solar_assumptions:
        assumption_map = {
            "peak_sun_hours":   "C24",
            "system_efficiency": "C25",
            "panel_wattage":    "C26",
            "cost_per_kwp":     "C27",
            "subsidy_pct":      "C28",
            "degradation_rate": "C29",
            "lifespan_years":   "C30",
        }
        for key, cell_ref in assumption_map.items():
            if key in solar_assumptions:
                ws[cell_ref] = solar_assumptions[key]
    
    wb.save(out_path)
    return out_path


def get_formula_results(output_path: Path) -> dict:
    """
    Read the calculated cells (after formulas are written).
    Returns a summary dict for display.
    NOTE: openpyxl can't evaluate formulas; values shown are formula strings.
    For actual computed values, user must open in Excel/LibreOffice.
    """
    wb = load_workbook(output_path, data_only=True)
    ws = wb.active
    
    results = {
        "consumer_name":      ws["C5"].value,
        "billing_month":      ws["C8"].value,
        "units_consumed":     ws["C13"].value,
        "daily_energy":       ws["C33"].value,
        "required_capacity":  ws["C34"].value,
        "recommended_size":   ws["C35"].value,
        "num_panels":         ws["C36"].value,
        "rooftop_area":       ws["C37"].value,
        "annual_generation":  ws["C38"].value,
        "gross_cost":         ws["C41"].value,
        "subsidy_amount":     ws["C42"].value,
        "net_cost":           ws["C43"].value,
        "annual_savings":     ws["C44"].value,
        "payback_period":     ws["C45"].value,
        "total_25yr_savings": ws["C46"].value,
        "co2_offset":         ws["C47"].value,
    }
    wb.close()
    return results


if __name__ == "__main__":
    # Quick test
    test_data = {
        "consumer_name": "Test Customer",
        "consumer_number": "123456789",
        "billing_month": "March 2024",
        "units_consumed": 450,
        "sanctioned_load": 5.0,
        "connected_load": 4.5,
        "tariff_category": "LT-I Residential",
        "per_unit_rate": 9.5,
        "fixed_charges": 200,
        "total_monthly_bill": 4475,
    }
    out = fill_excel(test_data)
    print(f"Generated: {out}")
